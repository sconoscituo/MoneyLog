"""엑셀 내보내기/가져오기 API 라우터"""
import io
import logging
from datetime import date, datetime
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy import select, extract
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.database import get_db
from app.models.expense import Expense
from app.models.category import Category

router = APIRouter(tags=["내보내기/가져오기"])
logger = logging.getLogger(__name__)

# 엑셀 헤더 색상
_HEADER_FILL = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_THIN_BORDER = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)


@router.get("/api/exports/excel")
async def export_excel(
    year: Optional[int] = Query(None, description="연도 필터"),
    month: Optional[int] = Query(None, description="월 필터 (1-12)"),
    db: AsyncSession = Depends(get_db),
):
    """지출 데이터를 Excel 파일로 다운로드한다."""
    try:
        stmt = (
            select(Expense)
            .options(selectinload(Expense.category))
            .order_by(Expense.expense_date, Expense.id)
        )
        if year:
            stmt = stmt.where(extract("year", Expense.expense_date) == year)
        if month:
            stmt = stmt.where(extract("month", Expense.expense_date) == month)

        result = await db.execute(stmt)
        expenses = result.scalars().all()

        # 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "지출 내역"

        # 헤더
        headers = ["번호", "날짜", "카테고리", "메모", "금액(원)", "비고"]
        col_widths = [8, 14, 14, 30, 16, 20]

        for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _THIN_BORDER
            ws.column_dimensions[cell.column_letter].width = width

        ws.row_dimensions[1].height = 22

        # 데이터 행
        for row_idx, exp in enumerate(expenses, start=2):
            row_data = [
                row_idx - 1,
                str(exp.expense_date),
                f"{exp.category.icon} {exp.category.name}",
                exp.memo or "",
                exp.amount,
                exp.note or "",
            ]
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = _THIN_BORDER
                cell.alignment = Alignment(vertical="center")
                if col_idx == 5:  # 금액 열 우측 정렬
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.number_format = '#,##0'

        # 합계 행
        total_row = len(expenses) + 2
        ws.cell(row=total_row, column=4, value="합계").font = Font(bold=True)
        total_cell = ws.cell(row=total_row, column=5, value=sum(e.amount for e in expenses))
        total_cell.font = Font(bold=True)
        total_cell.number_format = '#,##0'
        total_cell.alignment = Alignment(horizontal="right", vertical="center")

        # 파일명
        label = f"{year}년{month}월_" if year and month else ""
        filename = f"MoneyLog_{label}지출내역.xlsx"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
        )
    except Exception as e:
        logger.error(f"엑셀 내보내기 오류: {e}")
        raise HTTPException(status_code=500, detail="엑셀 파일 생성에 실패했습니다.")


@router.post("/api/imports/excel", status_code=201)
async def import_excel(
    file: UploadFile = File(..., description="지출 내역 Excel 파일 (.xlsx)"),
    db: AsyncSession = Depends(get_db),
):
    """Excel 파일을 업로드하여 지출을 일괄 등록한다.

    엑셀 형식: 날짜(YYYY-MM-DD), 카테고리명, 메모, 금액(숫자)
    첫 번째 행은 헤더로 간주하여 건너뜀.
    """
    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail=".xlsx 파일만 업로드할 수 있습니다.")

    try:
        contents = await file.read()
        wb = load_workbook(filename=io.BytesIO(contents), read_only=True, data_only=True)
        ws = wb.active

        # 카테고리 이름 → ID 매핑 (대소문자 무관, 이모지 포함 제거 후 비교)
        cats_result = await db.execute(select(Category).order_by(Category.id))
        categories = cats_result.scalars().all()
        cat_map: dict[str, int] = {c.name.strip(): c.id for c in categories}
        default_cat_id = categories[0].id if categories else 1

        added = 0
        errors: list[str] = []

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row_num, row in enumerate(rows, start=2):
            # 빈 행 건너뜀
            if not row or all(v is None for v in row):
                continue

            # 열 파싱: 날짜, 카테고리, 메모, 금액
            try:
                raw_date = row[0]
                raw_cat = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                raw_memo = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                raw_amount = row[3] if len(row) > 3 else None

                # 날짜 파싱
                if isinstance(raw_date, datetime):
                    expense_date = raw_date.date()
                elif isinstance(raw_date, date):
                    expense_date = raw_date
                elif isinstance(raw_date, str):
                    expense_date = date.fromisoformat(raw_date.strip()[:10])
                else:
                    errors.append(f"{row_num}행: 날짜 형식 오류 ({raw_date})")
                    continue

                # 금액 파싱
                amount = float(raw_amount)
                if amount <= 0:
                    errors.append(f"{row_num}행: 금액은 0보다 커야 합니다.")
                    continue

                # 카테고리 매핑 (없으면 기본값)
                # 이모지 앞 글자 제거 후 이름만 추출
                cat_name_clean = raw_cat.split(" ")[-1] if " " in raw_cat else raw_cat
                category_id = cat_map.get(raw_cat) or cat_map.get(cat_name_clean) or default_cat_id

                expense = Expense(
                    amount=amount,
                    category_id=category_id,
                    memo=raw_memo,
                    expense_date=expense_date,
                )
                db.add(expense)
                added += 1

            except (ValueError, TypeError) as parse_err:
                errors.append(f"{row_num}행: 파싱 오류 — {parse_err}")
                continue

        await db.flush()
        logger.info(f"엑셀 가져오기 완료: {added}건 등록, {len(errors)}건 오류")

        return {
            "message": f"{added}건의 지출이 등록되었습니다.",
            "added": added,
            "errors": errors,
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"엑셀 가져오기 오류: {e}")
        raise HTTPException(status_code=500, detail="엑셀 파일 처리에 실패했습니다.")
